import openpyxl


def write(a, b, c):
    # 打开已有的 .xlsx
    data = openpyxl.load_workbook('HKR.xlsx')  # 可读可写
    s = data["login_success"]
    s.cell(a, b, c)  # 行，列，值（行和列从1开始，不是从0开始）
    data.save("HKR.xlsx")  # 保存


def write1(a, b, c):
    # 打开已有的 .xlsx
    data = openpyxl.load_workbook('HKR.xlsx')  # 可读可写
    s = data["error_data"]
    s.cell(a, b, c)  # 行，列，值（行和列从1开始，不是从0开始）
    data.save("HKR.xlsx")  # 保存
